"""Export xlsx d'une collection (V0.9.0-gamma.2).

Une feuille avec :
- métadonnées de la collection en haut (cote, titre, type, fonds
  parent ou fonds représentés) ;
- une ligne d'entête avec les colonnes ;
- une ligne par item avec ses valeurs.

Format facile à éditer dans Excel/LibreOffice pour catalogage.
"""

from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from archives_tool.exporters._commun import composer_export
from archives_tool.exporters.mapping_dc import extraire_valeur
from archives_tool.exporters.rapport import RapportExport
from archives_tool.models import Collection, TypeCollection

# Colonnes affichées par item, dans l'ordre. Adaptées au workflow
# de catalogage : cote/titre/contexte d'abord, puis état et dates,
# puis identifiants externes.
_COLONNES = [
    ("cote", "Cote"),
    ("titre", "Titre"),
    ("fonds_cote", "Fonds"),
    ("etat_catalogage", "État"),
    ("date", "Date"),
    ("annee", "Année"),
    ("type_coar", "Type"),
    ("langue", "Langue"),
    ("description", "Description"),
    ("notes_internes", "Notes internes"),
    ("doi_nakala", "DOI Nakala"),
    ("nb_fichiers", "Nb fichiers"),
]

# Excel limite les noms de feuille à 31 caractères et interdit certains
# caractères (`[]:*?/\`).
_TITRE_FEUILLE_MAX = 31


def _slug_feuille(titre: str) -> str:
    interdits = set('[]:*?/\\')
    nettoye = "".join(c for c in titre if c not in interdits)
    return nettoye[:_TITRE_FEUILLE_MAX] or "Export"


def exporter_excel(
    session: Session,
    collection: Collection,
    chemin_sortie: Path,
) -> RapportExport:
    """Exporte une collection en xlsx pour catalogage manuel."""
    debut = time.monotonic()
    export = composer_export(session, collection)

    rapport = RapportExport(
        format="xlsx",
        chemin_sortie=chemin_sortie,
        nb_items_selectionnes=len(export.items),
        nb_fichiers_selectionnes=sum(len(ipe.item.fichiers) for ipe in export.items),
    )

    chemin_sortie.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = _slug_feuille(collection.titre)

    # Bandeau métadonnées de collection (lignes 1-4).
    type_libelle = (
        "miroir"
        if collection.type_collection == TypeCollection.MIROIR.value
        else ("transversale" if collection.fonds_id is None else "libre")
    )
    ws["A1"] = f"Collection : {collection.titre}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Cote : {collection.cote}"
    ws["A3"] = f"Type : {type_libelle}"
    if export.fonds_parent is not None:
        ws["A4"] = f"Fonds parent : {export.fonds_parent.titre} ({export.fonds_parent.cote})"
    elif export.fonds_representes:
        cotes = ", ".join(f.cote for f in export.fonds_representes)
        ws["A4"] = f"Fonds représentés : {cotes}"

    # Entêtes (ligne 6).
    en_tete_row = 6
    for col_idx, (_, libelle) in enumerate(_COLONNES, start=1):
        cell = ws.cell(row=en_tete_row, column=col_idx, value=libelle)
        cell.font = Font(bold=True)

    # Items (ligne 7+).
    for row_offset, ipe in enumerate(export.items, start=1):
        item = ipe.item
        row = en_tete_row + row_offset
        for col_idx, (champ, _) in enumerate(_COLONNES, start=1):
            if champ == "fonds_cote":
                valeur = ipe.fonds_cote
            elif champ == "nb_fichiers":
                valeur = len(item.fichiers)
            else:
                valeur = extraire_valeur(item, champ)
                if isinstance(valeur, list):
                    valeur = " | ".join(str(v) for v in valeur)
            ws.cell(row=row, column=col_idx, value=valeur)

    # Largeurs raisonnables sur les colonnes textuelles.
    largeurs = {"A": 18, "B": 40, "C": 10, "I": 40, "J": 30}
    for lettre, largeur in largeurs.items():
        ws.column_dimensions[lettre].width = largeur

    wb.save(chemin_sortie)
    rapport.duree_secondes = time.monotonic() - debut
    return rapport
