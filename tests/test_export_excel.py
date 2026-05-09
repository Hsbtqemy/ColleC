"""Tests de l'export xlsx (V0.9.0-gamma.2)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from archives_tool.api.services.collections import lire_collection_par_cote
from archives_tool.api.services.fonds import lire_fonds_par_cote
from archives_tool.exporters.excel import exporter_excel

# Fixture `session_avec_export` partagée définie dans tests/conftest.py.


def test_export_xlsx_structure_basique(
    session_avec_export: Session, tmp_path: Path
) -> None:
    """Bandeau métadonnées en haut, en-têtes en ligne 6, items à partir
    de la ligne 7."""
    fonds = lire_fonds_par_cote(session_avec_export, "HK")
    miroir = lire_collection_par_cote(session_avec_export, "HK", fonds_id=fonds.id)

    sortie = tmp_path / "hk.xlsx"
    rapport = exporter_excel(session_avec_export, miroir, sortie)
    assert sortie.is_file()
    assert rapport.nb_items_selectionnes == 3

    wb = load_workbook(sortie)
    ws = wb.active
    # Ligne 1 : titre de collection.
    assert "Hara-Kiri" in ws["A1"].value
    assert ws["A2"].value == "Cote : HK"
    assert ws["A3"].value == "Type : miroir"
    assert "HK" in ws["A4"].value  # fonds parent

    # Ligne 6 : entêtes.
    assert ws["A6"].value == "Cote"
    assert ws["B6"].value == "Titre"

    # Ligne 7+ : items.
    cotes = sorted(ws.cell(row=r, column=1).value for r in (7, 8, 9))
    assert cotes == ["HK-001", "HK-002", "HK-003"]


def test_export_xlsx_transversale_bandeau(
    session_avec_export: Session, tmp_path: Path
) -> None:
    """Pour une transversale, le bandeau liste les fonds représentés
    et la colonne Fonds varie d'un item à l'autre."""
    transv = lire_collection_par_cote(session_avec_export, "TRANSV")

    sortie = tmp_path / "transv.xlsx"
    exporter_excel(session_avec_export, transv, sortie)

    wb = load_workbook(sortie)
    ws = wb.active
    assert ws["A3"].value == "Type : transversale"
    assert "FA" in ws["A4"].value
    assert "HK" in ws["A4"].value

    # Items : 2 fonds différents en colonne C (Fonds).
    fonds_par_ligne = {ws.cell(row=r, column=3).value for r in (7, 8)}
    assert fonds_par_ligne == {"HK", "FA"}


def test_export_xlsx_titre_feuille_tronque(
    session_avec_export: Session, tmp_path: Path
) -> None:
    """Excel limite les noms de feuille à 31 caractères ; le slug
    tronque."""
    fonds = lire_fonds_par_cote(session_avec_export, "HK")
    miroir = lire_collection_par_cote(session_avec_export, "HK", fonds_id=fonds.id)
    miroir.titre = "Un titre extrêmement long qui dépasse largement la limite Excel"
    session_avec_export.commit()

    sortie = tmp_path / "long.xlsx"
    exporter_excel(session_avec_export, miroir, sortie)

    wb = load_workbook(sortie)
    assert len(wb.active.title) <= 31
