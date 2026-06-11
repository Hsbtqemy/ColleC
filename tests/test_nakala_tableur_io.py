"""Tests des écrivains tableur Nakala (Lot 1, T1.3) — CSV + xlsx."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from archives_tool.external.nakala.tableur import TableurNakala
from archives_tool.external.nakala.tableur_io import ecrire_csv, ecrire_xlsx


def _tableur() -> TableurNakala:
    return TableurNakala(
        colonnes=["identifier", "nkl:title", "dcterms:subject"],
        lignes=[
            {"identifier": "d1", "nkl:title": "Titre A", "dcterms:subject": "x | y"},
            {"identifier": "d2", "nkl:title": "Avec ; point-virgule", "dcterms:subject": ""},
        ],
    )


def test_csv_separateur_point_virgule_et_relecture(tmp_path: Path) -> None:
    chemin = tmp_path / "out.csv"
    ecrire_csv(_tableur(), chemin)
    # Relecture avec le même séparateur → mêmes lignes.
    with open(chemin, encoding="utf-8-sig", newline="") as f:
        lignes = list(csv.DictReader(f, delimiter=";"))
    assert [ligne["identifier"] for ligne in lignes] == ["d1", "d2"]
    assert lignes[0]["nkl:title"] == "Titre A"
    # Le champ contenant ';' est correctement quoté → pas de colonne en trop.
    assert lignes[1]["nkl:title"] == "Avec ; point-virgule"


def test_csv_separateur_configurable(tmp_path: Path) -> None:
    chemin = tmp_path / "out.csv"
    ecrire_csv(_tableur(), chemin, sep=",")
    texte = chemin.read_text(encoding="utf-8-sig")
    assert "identifier,nkl:title,dcterms:subject" in texte.splitlines()[0]


def test_csv_encodage_utf8_bom(tmp_path: Path) -> None:
    chemin = tmp_path / "out.csv"
    ecrire_csv(_tableur(), chemin)
    brut = chemin.read_bytes()
    assert brut.startswith(b"\xef\xbb\xbf")  # BOM utf-8-sig (Excel FR)


def test_xlsx_entetes_et_nb_lignes(tmp_path: Path) -> None:
    chemin = tmp_path / "out.xlsx"
    ecrire_xlsx(_tableur(), chemin, titre_collection="Ma collection")
    wb = load_workbook(chemin, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # Ligne 1 = bandeau titre collection, ligne 2 = entêtes, puis 2 données.
    assert any("Ma collection" in str(c) for c in rows[0] if c)
    entetes = [c for c in rows[1] if c is not None]
    assert entetes[:3] == ["identifier", "nkl:title", "dcterms:subject"]
    data = [r for r in rows[2:] if any(c is not None for c in r)]
    assert len(data) == 2
